"""
Обработчики управления заказами для администратора.
"""

from aiogram import Router, F, Bot
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile
)
from datetime import datetime
from decimal import Decimal
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from backend.app.config import settings

router = Router()


async def get_db():
    """Получает экземпляр базы данных."""
    from backend.app.services.database import DatabaseService
    
    db = DatabaseService(db_path=settings.DATABASE_PATH)
    await db.connect()
    return db


def is_admin(user_id: int) -> bool:
    """Проверяет, является ли пользователь администратором."""
    admin_ids_str = os.getenv("ADMIN_IDS", "") or getattr(settings, "ADMIN_IDS", "")
    
    if admin_ids_str:
        try:
            admin_ids = [int(id.strip()) for id in admin_ids_str.split(",") if id.strip().isdigit()]
            return user_id in admin_ids
        except (ValueError, AttributeError):
            pass
    
    return True  # Временно разрешаем всем для разработки


async def show_orders_menu(callback: CallbackQuery, bot: Bot):
    """Показывает главное меню управления заказами."""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ У вас нет прав администратора.", show_alert=True)
        return
    
    try:
        db = await get_db()
        
        # Получаем статистику заказов
        total_revenue = await db.fetch_one(
            "SELECT COALESCE(SUM(total_amount), 0) as total FROM orders"
        )
        avg_order = await db.fetch_one(
            "SELECT COALESCE(AVG(total_amount), 0) as avg FROM orders"
        )
        orders_count = await db.fetch_one(
            "SELECT COUNT(*) as cnt FROM orders"
        )
        
        await db.disconnect()
        
        stats = {
            "total_revenue": float(total_revenue["total"]) if total_revenue and isinstance(total_revenue["total"], Decimal) else (total_revenue["total"] if total_revenue else 0),
            "average_order": float(avg_order["avg"]) if avg_order and isinstance(avg_order["avg"], Decimal) else (avg_order["avg"] if avg_order else 0),
            "orders_count": orders_count["cnt"] if orders_count else 0
        }
        
        menu_text = f"""
<b>📋 Управление заказами</b>

<b>Статистика:</b>
📊 Всего заказов: {stats.get('orders_count', 0)}
💰 Общая выручка: {stats.get('total_revenue', 0):.2f} ₽
📈 Средний чек: {stats.get('average_order', 0):.2f} ₽

Выберите действие:
"""
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 Все заказы", callback_data="admin_orders_list_all")],
            [InlineKeyboardButton(text="⏳ Ожидают", callback_data="admin_orders_list_pending")],
            [InlineKeyboardButton(text="✅ Подтверждены", callback_data="admin_orders_list_confirmed")],
            [InlineKeyboardButton(text="📦 Доставляются", callback_data="admin_orders_list_shipped")],
            [InlineKeyboardButton(text="✓ Доставлены", callback_data="admin_orders_list_delivered")],
            [InlineKeyboardButton(text="❌ Отменены", callback_data="admin_orders_list_cancelled")],
            [InlineKeyboardButton(text="📊 Статистика", callback_data="admin_orders_statistics")],
            [InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="admin_orders_export")],
            [InlineKeyboardButton(text="◀️ Назад", callback_data="admin_back_to_menu")]
        ])
        
        try:
            await callback.message.edit_text(menu_text, reply_markup=keyboard)
        except Exception:
            await callback.message.answer(menu_text, reply_markup=keyboard)
        await callback.answer()
        
    except Exception as e:
        print(f"Error showing orders menu: {e}")
        import traceback
        traceback.print_exc()
        await callback.answer("❌ Ошибка при загрузке меню заказов.", show_alert=True)


async def show_orders_list(callback: CallbackQuery, bot: Bot, status: str = None, page: int = 0):
    """Показывает список заказов с фильтрами и пагинацией."""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ У вас нет прав администратора.", show_alert=True)
        return
    
    try:
        db = await get_db()
        
        # Формируем условия фильтрации
        conditions = []
        params = []
        
        if status:
            conditions.append("o.status = ?")
            params.append(status)
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        limit = 10
        offset = page * limit
        
        orders = await db.fetch_all(
            f"""SELECT o.*, 
                      s.name as shop_name,
                      u.telegram_id as user_telegram_id,
                      u.username as user_username,
                      u.first_name as user_first_name,
                      u.last_name as user_last_name
               FROM orders o
               LEFT JOIN shops s ON o.shop_id = s.id
               LEFT JOIN users u ON o.user_id = u.id
               WHERE {where_clause}
               ORDER BY o.created_at DESC
               LIMIT ? OFFSET ?""",
            tuple(params + [limit, offset])
        )
        
        await db.disconnect()
        
        # Преобразуем Decimal в float
        orders_list = []
        for order in orders:
            order_dict = dict(order)
            if order_dict.get("total_amount") is not None:
                if isinstance(order_dict["total_amount"], Decimal):
                    order_dict["total_amount"] = float(order_dict["total_amount"])
            orders_list.append(order_dict)
        
        orders = orders_list
        
        status_names = {
            None: "Все заказы",
            "pending": "Ожидают",
            "confirmed": "Подтверждены",
            "processing": "Обрабатываются",
            "shipped": "Доставляются",
            "delivered": "Доставлены",
            "cancelled": "Отменены"
        }
        
        if not orders:
            text = f"<b>📋 {status_names.get(status, 'Заказы')}</b>\n\nЗаказов не найдено."
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="◀️ Назад", callback_data="admin_orders_menu")]
            ])
            await callback.message.edit_text(text, reply_markup=keyboard)
            await callback.answer()
            return
        
        text = f"<b>📋 {status_names.get(status, 'Заказы')}</b>\n\n"
        keyboard_buttons = []
        
        for order in orders:
            status_emoji = {
                "pending": "⏳",
                "confirmed": "✅",
                "processing": "🔄",
                "shipped": "📦",
                "delivered": "✓",
                "cancelled": "❌"
            }.get(order.get("status"), "📋")
            
            created_at = order.get("created_at", "")
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    created_at = dt.strftime("%d.%m.%Y %H:%M")
                except:
                    pass
            
            text += f"{status_emoji} <b>#{order['id']}</b> - {order.get('order_number', 'N/A')}\n"
            text += f"   Магазин: {order.get('shop_name', 'Неизвестно')}\n"
            text += f"   Клиент: {order.get('user_first_name', '')} {order.get('user_last_name', '')}\n"
            text += f"   Сумма: {order.get('total_amount', 0):.2f} ₽\n"
            text += f"   Дата: {created_at}\n\n"
            
            keyboard_buttons.append([
                InlineKeyboardButton(
                    text=f"#{order['id']} - {order.get('order_number', 'N/A')}",
                    callback_data=f"admin_order_view_{order['id']}"
                )
            ])
        
        # Пагинация
        nav_buttons = []
        if page > 0:
            nav_buttons.append(
                InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_orders_list_{status or 'all'}_{page-1}")
            )
        
        if len(orders) == 10:  # Если получили полную страницу, есть еще заказы
            nav_buttons.append(
                InlineKeyboardButton(text="Вперед ▶️", callback_data=f"admin_orders_list_{status or 'all'}_{page+1}")
            )
        
        if nav_buttons:
            keyboard_buttons.append(nav_buttons)
        
        keyboard_buttons.append([
            InlineKeyboardButton(text="◀️ Назад к меню", callback_data="admin_orders_menu")
        ])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
        
        try:
            await callback.message.edit_text(text, reply_markup=keyboard)
        except Exception:
            await callback.message.answer(text, reply_markup=keyboard)
        await callback.answer()
        
    except Exception as e:
        print(f"Error showing orders list: {e}")
        import traceback
        traceback.print_exc()
        try:
            await callback.answer("❌ Ошибка при загрузке списка заказов.", show_alert=True)
        except:
            pass


async def show_order_details(callback: CallbackQuery, bot: Bot, order_id: int):
    """Показывает детали заказа."""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ У вас нет прав администратора.", show_alert=True)
        return
    
    try:
        db = await get_db()
        
        order = await db.fetch_one(
            """SELECT o.*, 
                      s.name as shop_name,
                      u.telegram_id as user_telegram_id,
                      u.username as user_username,
                      u.first_name as user_first_name,
                      u.last_name as user_last_name
               FROM orders o
               LEFT JOIN shops s ON o.shop_id = s.id
               LEFT JOIN users u ON o.user_id = u.id
               WHERE o.id = ?""",
            (order_id,)
        )
        
        if not order:
            await db.disconnect()
            await callback.answer("❌ Заказ не найден", show_alert=True)
            return
        
        # Получаем товары заказа
        items = await db.fetch_all(
            """SELECT oi.*, 
                      COALESCE(oi.product_name, p.name, 'Товар удалён') as product_name
               FROM order_items oi
               LEFT JOIN products p ON oi.product_id = p.id
               WHERE oi.order_id = ?""",
            (order_id,)
        )
        
        await db.disconnect()
        
        # Преобразуем Decimal в float
        order_dict = dict(order)
        if order_dict.get("total_amount") is not None:
            if isinstance(order_dict["total_amount"], Decimal):
                order_dict["total_amount"] = float(order_dict["total_amount"])
        if order_dict.get("delivery_fee") is not None:
            if isinstance(order_dict["delivery_fee"], Decimal):
                order_dict["delivery_fee"] = float(order_dict["delivery_fee"])
        if order_dict.get("promo_discount_amount") is not None:
            if isinstance(order_dict["promo_discount_amount"], Decimal):
                order_dict["promo_discount_amount"] = float(order_dict["promo_discount_amount"])
        
        order_dict["items"] = []
        for item in items:
            item_dict = dict(item)
            if item_dict.get("price") is not None:
                if isinstance(item_dict["price"], Decimal):
                    item_dict["price"] = float(item_dict["price"])
            order_dict["items"].append(item_dict)
        
        order = order_dict
        
        status_emoji = {
            "pending": "⏳",
            "confirmed": "✅",
            "processing": "🔄",
            "shipped": "📦",
            "delivered": "✓",
            "cancelled": "❌"
        }.get(order.get("status"), "📋")
        
        created_at = order.get("created_at", "")
        if created_at:
            try:
                dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                created_at = dt.strftime("%d.%m.%Y %H:%M")
            except:
                pass
        
        text = f"""
<b>{status_emoji} Заказ #{order_id}</b>

<b>Номер заказа:</b> {order.get('order_number', 'N/A')}
<b>Статус:</b> {order.get('status', 'Неизвестно')}
<b>Дата создания:</b> {created_at}

<b>Магазин:</b> {order.get('shop_name', 'Неизвестно')}

<b>Клиент:</b>
👤 {order.get('user_first_name', '')} {order.get('user_last_name', '')}
📱 Telegram: @{order.get('user_username', 'не указан')}
ID: {order.get('user_telegram_id', 'не указан')}

<b>Адрес доставки:</b> {order.get('delivery_address', 'Не указан')}
<b>Тип доставки:</b> {order.get('delivery_type', 'Не указан')}
<b>Комментарий:</b> {order.get('delivery_comment', 'Нет')}

<b>Товары:</b>
"""
        
        items = order.get("items", [])
        for item in items:
            text += f"• {item.get('product_name', 'Товар')} x{item.get('quantity', 1)} = {item.get('price', 0) * item.get('quantity', 1):.2f} ₽\n"
        
        text += f"""
<b>Итого:</b>
💰 Сумма товаров: {order.get('total_amount', 0) - order.get('delivery_fee', 0) - order.get('promo_discount_amount', 0):.2f} ₽
🚚 Доставка: {order.get('delivery_fee', 0):.2f} ₽
🎫 Скидка по промокоду: {order.get('promo_discount_amount', 0):.2f} ₽
<b>💵 К оплате: {order.get('total_amount', 0):.2f} ₽</b>
"""
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="◀️ Назад к списку", callback_data="admin_orders_list_all_0")]
        ])
        
        try:
            await callback.message.edit_text(text, reply_markup=keyboard)
        except Exception:
            await callback.message.answer(text, reply_markup=keyboard)
        await callback.answer()
        
    except Exception as e:
        print(f"Error showing order details: {e}")
        import traceback
        traceback.print_exc()
        try:
            await callback.answer("❌ Ошибка при загрузке заказа.", show_alert=True)
        except:
            pass


async def export_orders_to_excel(callback: CallbackQuery, bot: Bot):
    """Экспортирует заказы в Excel файл."""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ У вас нет прав администратора.", show_alert=True)
        return
    
    try:
        await callback.answer("📥 Генерация файла...")
        
        # Получаем все заказы напрямую из БД
        db = await get_db()
        
        all_orders = await db.fetch_all(
            """SELECT o.*, 
                      s.name as shop_name,
                      u.telegram_id as user_telegram_id,
                      u.username as user_username,
                      u.first_name as user_first_name,
                      u.last_name as user_last_name
               FROM orders o
               LEFT JOIN shops s ON o.shop_id = s.id
               LEFT JOIN users u ON o.user_id = u.id
               ORDER BY o.created_at DESC"""
        )
        
        await db.disconnect()
        
        # Преобразуем Decimal в float
        orders_list = []
        for order in all_orders:
            order_dict = dict(order)
            if order_dict.get("total_amount") is not None:
                if isinstance(order_dict["total_amount"], Decimal):
                    order_dict["total_amount"] = float(order_dict["total_amount"])
            if order_dict.get("delivery_fee") is not None:
                if isinstance(order_dict["delivery_fee"], Decimal):
                    order_dict["delivery_fee"] = float(order_dict["delivery_fee"])
            if order_dict.get("promo_discount_amount") is not None:
                if isinstance(order_dict["promo_discount_amount"], Decimal):
                    order_dict["promo_discount_amount"] = float(order_dict["promo_discount_amount"])
            orders_list.append(order_dict)
        
        all_orders = orders_list
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        # Заголовки
        headers = [
            "ID", "Номер заказа", "Дата", "Статус", "Магазин", 
            "Клиент", "Телефон", "Адрес", "Тип доставки",
            "Сумма товаров", "Доставка", "Скидка", "Итого"
        ]
        
        # Стиль для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Данные
        for row_num, order in enumerate(all_orders, 2):
            created_at = order.get("created_at", "")
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    created_at = dt.strftime("%d.%m.%Y %H:%M")
                except:
                    pass
            
            subtotal = order.get('total_amount', 0) - order.get('delivery_fee', 0) - order.get('promo_discount_amount', 0)
            
            row_data = [
                order.get('id', ''),
                order.get('order_number', ''),
                created_at,
                order.get('status', ''),
                order.get('shop_name', ''),
                f"{order.get('user_first_name', '')} {order.get('user_last_name', '')}".strip(),
                order.get('phone', ''),
                order.get('delivery_address', ''),
                order.get('delivery_type', ''),
                subtotal,
                order.get('delivery_fee', 0),
                order.get('promo_discount_amount', 0),
                order.get('total_amount', 0)
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        # Отправляем файл
        file = FSInputFile(tmp_path, filename=f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        await bot.send_document(
            chat_id=callback.message.chat.id,
            document=file,
            caption=f"📥 Экспорт заказов\n\nВсего заказов: {len(all_orders)}"
        )
        
        # Удаляем временный файл
        os.unlink(tmp_path)
        
        await callback.answer("✅ Файл отправлен")
        
    except Exception as e:
        print(f"Error exporting orders: {e}")
        import traceback
        traceback.print_exc()
        await callback.answer("❌ Ошибка при экспорте заказов.", show_alert=True)


async def show_orders_statistics(callback: CallbackQuery, bot: Bot):
    """Показывает статистику по заказам."""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ У вас нет прав администратора.", show_alert=True)
        return
    
    try:
        db = await get_db()
        
        # Общая выручка
        total_revenue = await db.fetch_one(
            "SELECT COALESCE(SUM(total_amount), 0) as total FROM orders"
        )
        
        # Средний чек
        avg_order = await db.fetch_one(
            "SELECT COALESCE(AVG(total_amount), 0) as avg FROM orders"
        )
        
        # Количество заказов
        orders_count = await db.fetch_one(
            "SELECT COUNT(*) as cnt FROM orders"
        )
        
        # Заказы по статусам
        orders_by_status = await db.fetch_all(
            """SELECT status, COUNT(*) as cnt 
               FROM orders 
               GROUP BY status"""
        )
        
        await db.disconnect()
        
        stats = {
            "total_revenue": float(total_revenue["total"]) if total_revenue and isinstance(total_revenue["total"], Decimal) else (total_revenue["total"] if total_revenue else 0),
            "average_order": float(avg_order["avg"]) if avg_order and isinstance(avg_order["avg"], Decimal) else (avg_order["avg"] if avg_order else 0),
            "orders_count": orders_count["cnt"] if orders_count else 0,
            "orders_by_status": {row["status"]: row["cnt"] for row in orders_by_status}
        }
        
        text = f"""
<b>📊 Статистика по заказам</b>

<b>Общая статистика:</b>
📋 Всего заказов: {stats.get('orders_count', 0)}
💰 Общая выручка: {stats.get('total_revenue', 0):.2f} ₽
📈 Средний чек: {stats.get('average_order', 0):.2f} ₽

<b>Заказы по статусам:</b>
"""
        
        orders_by_status = stats.get('orders_by_status', {})
        for status, count in orders_by_status.items():
            status_emoji = {
                "pending": "⏳",
                "confirmed": "✅",
                "processing": "🔄",
                "shipped": "📦",
                "delivered": "✓",
                "cancelled": "❌"
            }.get(status, "📋")
            text += f"{status_emoji} {status}: {count}\n"
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="◀️ Назад", callback_data="admin_orders_menu")]
        ])
        
        await callback.message.edit_text(text, reply_markup=keyboard)
        await callback.answer()
        
    except Exception as e:
        print(f"Error showing orders statistics: {e}")
        import traceback
        traceback.print_exc()
        await callback.answer("❌ Ошибка при загрузке статистики.", show_alert=True)


# Обработчики callback
@router.callback_query(F.data == "admin_orders_menu")
async def callback_orders_menu(callback: CallbackQuery, bot: Bot):
    """Обработчик кнопки меню заказов."""
    await show_orders_menu(callback, bot)


@router.callback_query(F.data.startswith("admin_orders_list_"))
async def callback_orders_list(callback: CallbackQuery, bot: Bot):
    """Обработчик списка заказов."""
    parts = callback.data.split("_")
    status = parts[3] if len(parts) > 3 and parts[3] != "all" else None
    page = int(parts[4]) if len(parts) > 4 else 0
    await show_orders_list(callback, bot, status, page)


@router.callback_query(F.data.startswith("admin_order_view_"))
async def callback_order_view(callback: CallbackQuery, bot: Bot):
    """Обработчик просмотра заказа."""
    order_id = int(callback.data.split("_")[3])
    await show_order_details(callback, bot, order_id)


@router.callback_query(F.data == "admin_orders_export")
async def callback_orders_export(callback: CallbackQuery, bot: Bot):
    """Обработчик экспорта заказов."""
    await export_orders_to_excel(callback, bot)


@router.callback_query(F.data == "admin_orders_statistics")
async def callback_orders_statistics(callback: CallbackQuery, bot: Bot):
    """Обработчик статистики заказов."""
    await show_orders_statistics(callback, bot)

